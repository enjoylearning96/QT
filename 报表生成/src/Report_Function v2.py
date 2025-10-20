import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

class ReportGenerator():
    
    def __init__(sellf,database,date,shift,filename):
        
        # 创建Excel写入器
        self.writer = pd.ExcelWriter(filename, engine='openpyxl')
        self.database=database
        self.date = date
        self.shift = shift
        self.init_data()
        self.create_front()
        self.create_back()
    
    def init_data(self):
        previous_day = (datetime.strptime(date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
        next_day = (datetime.strptime(date, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
        self.shifts_datas = {
            '二班':None,
            '三班':None,
            '一班':None,
            'date':None
            }
        if self.shift == '二班' or self.shift == '三班':
            self.shifts_datas['二班'] = self.database.get_shift_records(date=self.date, shift='二班')
            self.shifts_datas['三班'] = self.database.get_shift_records(date=self.date, shift='三班')
            self.shifts_datas['一班'] = self.database.get_shift_records(date=next_day, shift='一班')
            date_start = datetime.strptime(self.date, "%Y-%m-%d")
            date_end = datetime.strptime(next_day, "%Y-%m-%d")
            self.shifts_datas['date'] = f"{date_start.month}月{date_start.day}日二班至{date_end.month}月{date_end.day}日一班"
        else:
            self.shifts_datas['二班'] = self.database.get_shift_records(date=previous_day, shift='二班')
            self.shifts_datas['三班'] = self.database.get_shift_records(date=previous_day, shift='三班')
            self.shifts_datas['一班'] = self.database.get_shift_records(date=self.date, shift='一班')
            date_start = datetime.strptime(previous_day, "%Y-%m-%d")
            date_end = datetime.strptime(self.date, "%Y-%m-%d")
            self.shifts_datas['date'] = f"{date_start.month}月{date_start.day}日二班至{date_end.month}月{date_end.day}日一班"
        
    def create_front(self):
        shovels = self.database.get_vehicle_data(vehicle_type='电铲',vehicle_available=1)
        # 当前班次所使用的电铲数
        num=len(self.shifts_datas[self.shift])
        # 所有可使用电铲
        num_shovels = len(shovels)
        front_data = [None]*(3+num+1+1+num_shovels+1+5*num)
        front_data[0] = ["黑岱沟露天煤矿无人驾驶运行日报表"] + [""]*13
        front_data[1]=[f"{shifts_datas['date']}"]
        if self.shift == '一班':
            front_data[1].append(f"15 台无人驾驶卡车，{num}编组运行，故障  台，原车故障 0 台，无人故障  台，待令  台，调试 0 台，交付自营 0 台，4#395累计运行时间 0 小时，2#35累计运行时间 0 小时，单编组累计拉运 0 车，完成剥离量 0 立方米，平均运距 0 公里。")
        front_data[1].append(f"主要影响因素：")
        for shift_datas in self.shifts_datas:
            if  shift_datas:  
                front_data[1].append(f"{shift_datas[0]['shift']}：{shift_datas[0]['operating_effect_factor']};")
        front_data[2] = ["生产完成情况", "", "铲组", "", "日计划（万m³）", "日完成（万m³）", "日超欠（万m³）", "", "月计划（万m³）", "月完成（万m³）", "月超欠（万m³）", "", "完成月计划的百分比"]

        
        i=0
        if self.shift == '二班' or self.shift == '三班':
            shift_datas = self.shifts_datas['二班']
            for shift_data in self.shift_datas:
                front_data[3+i] = ["", "", f"{shift_data['shovel_id']}", "", f"{shift_data['daily_plan']}", "", "", "", f"{shift_data['monthly_plan']}", "", "", "", ""]
                i+=1
        else:
            shift_datas = self.shifts_datas['一班']
            for shift_data in self.shift_datas:
                front_data[3+i] = ["", "", f"{shift_data['shovel_id']}", "", f"{shift_data['daily_plan']}", "", "", "", f"{shift_data['monthly_plan']}", "", "", "", ""]
                i+=1        
        
        front_data[3+i]=["总计", "", "", "", "", "", "", "", "", "", "", "", ""]
        front_data[4+i]=["年度完成情况", "", "铲组", "", "日期", "", "年度累计运行时间 (小时）", "", "年度累计拉运车数（车）", "", "年度累计完成产量（m³）", "", ""]
        
        j=0
        for shovel in shovels:
            if shovel['vehicle_number'] == '"7#55"':                   
                front_data[5+i+j] = ["", "", "7#55", "", "","1月1日一班-6月28日一班", "", "1859.8", "", "25005", "", "2587802", "", ""]
            elif shovel['vehicle_number'] == '"8#55"':
                front_data[5+i+j] = ["", "", "8#55", "", "","5月16日二班-5月29日一班", "", "146.1", "", "1999", "", "174645", "", ""]
            elif shovel['vehicle_number'] == '"8#55"':
                front_data[5+i+j] = ["", "", "5#55", "","", "6月12日二班-6月30日一班", "", "124.2", "", "758", "", "79765", "", ""]
            else :
                front_data[5+i+j] = [""]*14
            j+=1
        front_data[5+i+j] = ["", "", "", f"{shift_datas[0]['foreman']}", "", "", f"{shift_datas[0]['foreman']}", "", "", "", f"{shift_datas[0]['foreman']}", "","", ""]
        k=0
        for shift_data in self.shift_datas:
            front_data[6+i+j+5*k] = [f"{shift_data['shovel_id']}", "配车数", f"{shift_data['vehicle_available_count']}", "台", f"{shift_data['operating_status']}", "", "0", "台", "", "", "0", "台", ""]
            front_data[7+i+j+5*k] = ["", "拉运车数", f"{shift_data['vehicle_count']}", "车", "", "", "0", "车", "", "", "0", "车", ""]
            front_data[8+i+j+5*k] = ["", "运行时长", f"{shift_data['operating_time']}", "小时", "", "", "0", "小时", "", "", "0", "小时", ""]
            front_data[9+i+j+5*k] = ["", "运距", f"{shift_data['operating_length']}", "公里", "", "", "0", "公里", "", "", "0", "公里", ""]
            front_data[10+i+j+5*k] = ["", "产量", f"{shift_data['production']}", "m³", "", "", "0", "m³", "", "", "0", "m³", ""]
            k+=1
            

    # ==================== 正面工作表 ====================
    '''
    front_data = [
        ["黑岱沟露天煤矿无人驾驶运行日报表"] + [""]*13,
        ["7月31日二班至8月01日一班，15 台无人驾驶卡车，未运行，故障  台，原车故障 0 台，无人故障  台，待令  台，调试 0 台，交付自营 0 台，4#395累计运行时间 0 小时，2#35累计运行时间 0 小时，单编组累计拉运 0 车，完成剥离量 0 立方米，平均运距 0 公里。主要影响因素：二班：下雨影响，集控室搬迁停电；"] + [""]*13,
        ["生产完成情况", "", "铲组", "", "日计划（万m³）", "日完成（万m³）", "日超欠（万m³）", "", "月计划（万m³）", "月完成（万m³）", "月超欠（万m³）", "", "完成月计划的百分比"],
        ["", "", "4#395", "", "0.903", "", "-0.903", "", "28", "17.463", "-10.537", "", "62.4%"],
        ["", "", "2#35", "", "0.903", "", "-0.903", "", "28", "21.804", "-6.196", "", "77.9%"],
        ["总计", "", "", "", "1.806", "39.267", "-1.806", "", "56", "39.267", "-16.733", "", "70.1%"],
        ["年度完成情况", "", "铲组", "", "日期", "", "年度累计运行时间 (小时）", "", "年度累计拉运车数（车）", "", "年度累计完成产量（m³）", "", ""],
        ["", "", "7#55", "", "1月1日一班-6月28日一班", "", "1859.8", "", "25005", "", "2587802", "", ""],
        ["", "", "4#395", "", "1月1日一班-8月01日一班", "", "2296.9", "", "23846", "", "1873535", "", ""],
        ["", "", "8#55", "", "5月16日二班-5月29日一班", "", "146.1", "", "1999", "", "174645", "", ""],
        ["", "", "2#35", "", "6月26日二班-8月01日一班", "", "262.8", "", "3209", "", "293671", "", ""],
        ["", "", "5#55", "", "6月12日二班-6月30日一班", "", "124.2", "", "758", "", "79765", "", ""],
        ["", "", "二班（刘淼）", "", "", "", "三班（刘淼）", "", "", "", "一班（刘淼）", "", ""],
        ["4#395编组", "配车数", "0", "台", "08:30-16:30 下雨影响，集控室搬迁停电", "", "0", "台", "16:30-00:30 雨后恢复道路", "", "0", "台", ""],
        ["", "拉运车数", "0", "车", "", "", "0", "车", "", "", "0", "车", ""],
        ["", "运行时长", "0", "小时", "", "", "0", "小时", "", "", "0", "小时", ""],
        ["", "运距", "0", "公里", "", "", "0", "公里", "", "", "0", "公里", ""],
        ["", "产量", "0", "m³", "", "", "0", "m³", "", "", "0", "m³", ""],
        ["2#35编组", "配车数", "0", "台", "08:30-16:30 下雨影响，集控室搬迁停电", "", "0", "台", "16:30-00:30 雨后恢复道路", "", "0", "台", ""],
        ["", "拉运车数", "0", "车", "", "", "0", "车", "", "", "0", "车", ""],
        ["", "运行时长", "0", "小时", "", "", "0", "小时", "", "", "0", "小时", ""],
        ["", "运距", "0", "公里", "", "", "0", "公里", "", "", "0", "公里", ""],
        ["", "产量", "0", "m³", "", "", "0", "m³", "", "", "0", "m³", ""],
        ["", "", "", "", "", "", "", "", "", "", "", "", ""]
    ]
    '''
    df_front = pd.DataFrame(front_data)
    df_front.to_excel(self.writer, sheet_name='正面', index=False, header=False)
    
    # ==================== 反面工作表 ====================
    back_data = [
        ["卡车类型", "编号", "二班", "", "", "", "", "", "三班", "", "", "", "", "", "", "一班", "", "", "", ""],
        ["", "", "运行时长", "接管次数", "故障名称", "故障类型", "处理情况", "故障时长", "运行时长", "接管次数", "故障名称", "故障类型", "处理情况", "故障时长", "运行时长", "接管次数", "故障名称", "故障类型", "处理情况", "故障时长"],
        ["930E", "501", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "502", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "503", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "504", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "506", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "507", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "508", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "509", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "510", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "511", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "512", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "513", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "514", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["830E", "402", "0", "", "", "", "", "", "0", "", "", "", "", "", "0", "", "", "", "", ""],
        ["", "405", "0", "", "定位异常", "无人驾驶系统故障", "待处理", "8", "0", "", "", "", "", "", "0", "", "", "", "", ""]
    ]
    
    df_back = pd.DataFrame(back_data)
    df_back.to_excel(writer, sheet_name='反面', index=False, header=False)
    
    # 保存Excel文件
    writer.close()
    
    # ==================== 精确样式设置 ====================
    wb = load_workbook(filename)
    
    # 设置正面工作表样式
    front_sheet = wb['正面']
    
    # 1. 标题样式
    # front_sheet.merge_cells('A1:N1')
    title_cell = front_sheet['A1']
    title_cell.font = Font(name='方正小标宋简体', size=30, bold=False, color='000000')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    # title_cell.fill = PatternFill("solid", fgColor="D9D9D9")
    
    # 2. 第二行描述文本
    # front_sheet.merge_cells('A2:N2')
    desc_cell = front_sheet['A2']
    desc_cell.font = Font(name='等线', size=20, bold=True,color='000000')
    desc_cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    # 3. 完成情况区域样式
    header_font = Font(name='等线', size=16, bold=True, color='000000')
    header_fill = PatternFill("solid", fgColor="FFF2CC")
    for row in front_sheet.iter_rows(min_row=3, max_row=12, min_col=1, max_col=14):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 4. 数据区域样式
    data_font = Font(name='等线', size=18, color='000000')
    for row in front_sheet.iter_rows(min_row=13, max_row=23, min_col=1, max_col=14):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    data_bianzu_fill=PatternFill("solid", fgColor="E2EFDA")
    for row in front_sheet.iter_rows(min_row=13, max_row=23, min_col=1, max_col=2):
        for cell in row:
            cell.fill = data_bianzu_fill
    
    # 5. 合并单元格
    merge_ranges = [
        'A1:N1',
        'A2:N2',
        'A3:B5',
        'C3:D3','G3:H3','K3:L3','M3:N3',
        'C4:D4','G4:H4','K4:L4','M4:N4',
        'C5:D5','G5:H5','K5:L5','M5:N5',
        'A6:D6','G6:H6','K6:L6','M6:N6',
        'A7:B12','C7:D7','E7:F7','G7:H7','I7:J7','K7:N7',
        'C8:D8','E8:F8','G8:H8','I8:J8','K8:N8',
        'C9:D9','E9:F9','G9:H9','I9:J9','K9:N9',
        'C10:D10','E10:F10','G10:H10','I10:J10','K10:N10',
        'C11:D11','E11:F11','G11:H11','I11:J11','K11:N11',
        'C12:D12','E12:F12','G12:H12','I12:J12','K12:N12',
        'A13:B13','C13:F13','G13:J13','K13:N13',
        'A14:A18','E14:F18','I14:J18','M14:N18',
        'A19:A23','E19:F23','I19:J23','M19:N23'
    ]
    for merge_range in merge_ranges:
        front_sheet.merge_cells(merge_range)
    
    # 6. 边框设置
    thin_border = Border(
        left=Side(style='thin', color='000000'), 
        right=Side(style='thin', color='000000'), 
        top=Side(style='thin', color='000000'), 
        bottom=Side(style='thin', color='000000')
    )
    for row in front_sheet.iter_rows(min_row=1, max_row=23, min_col=1, max_col=14):
        for cell in row:
            cell.border = thin_border
    
    # 7. 精确列宽设置（单位：字符宽度）
    col_widths = {
        'A': 18.71, 'B': 15, 
        'C': 15, 'D': 15, 'E': 42.86, 'F': 42.86, 
        'G': 15, 'H': 15, 'I': 42.86, 'J': 42.86, 
        'K': 15, 'L': 15, 'M': 42.86, 'N': 42.86
    }
    for col, width in col_widths.items():
        front_sheet.column_dimensions[col].width = width
    
    # 8. 行高设置
    front_sheet.row_dimensions[1].height = 70
    front_sheet.row_dimensions[2].height = 120
    for i in range(3, 13):
        front_sheet.row_dimensions[i].height = 40
    for i in range(13, 24):
        front_sheet.row_dimensions[i].height = 100
    
    # ==================== 反面工作表样式 ====================
    back_sheet = wb['反面']
    
    
    # 1. 数据区域样式
    data_font = Font(name='等线', size=12, color='000000')
    for row in back_sheet.iter_rows(min_row=1, max_row=17, min_col=1, max_col=20):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 2. 表头样式
    title_font = Font(name='等线', size=14, bold=False, color='000000')
    for row in back_sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=20):
        for cell in row:
            cell.font = title_font  
    
    # 4. 合并单元格
    merge_ranges = [
        'A1:A2','B1:B2','C1:H1','I1:N1','O1:T1',        
        'A3:A15',
        'A16:A17'
    ]
    for merge_range in merge_ranges:
        back_sheet.merge_cells(merge_range) 
    # 5. 边框设置
    for row in back_sheet.iter_rows(min_row=1, max_row=17, min_col=1, max_col=20):
        for cell in row:
            cell.border = thin_border
    
    # 6. 精确列宽设置
    col_widths = {
        'A': 12, 'B': 12, 
        'C': 12, 'D': 12, 'E': 29, 'F': 29, 'G': 12, 'H': 12,
        'I': 12, 'J': 12, 'K': 29, 'L': 29, 'M': 12, 'N': 12, 
        'O': 12, 'P': 12, 'Q': 29,'R': 12, 'S': 12, 'T': 12
    }
    for col, width in col_widths.items():
        back_sheet.column_dimensions[col].width = width
    
    # 7. 行高设置
    back_sheet.row_dimensions[1].height = 50
    back_sheet.row_dimensions[2].height = 100
    for i in range(3, 18):
        back_sheet.row_dimensions[i].height = 45
    
    # 保存文件
    wb.save(filename)
    wb.close()